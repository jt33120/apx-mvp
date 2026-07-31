"""Server-side office rendering for the pièce viewer (Story 3.5c-2, AD-14's render path).

Renders a ``.docx`` (mammoth) or ``.xlsx`` (openpyxl) pièce to **sanitised inline HTML**, inside the
tenant boundary — no pièce byte ever leaves for a third-party conversion service. The security spine
is one choke point: ``_rendered`` is the ONLY place a ``RenderedDocument`` is built, and it runs
every fragment through ``_sanitize`` (nh3, a strict allow-list) before the HTML can be returned. So
an **adversarial** document — a ``.docx``/``.xlsx`` from opposing counsel carrying ``<script>``, an
``onerror`` handler, a ``javascript:`` link or a tracking image — cannot execute or phone home when
the SPA displays it (a structural property seals it, Story 3.5c-2).

mammoth / openpyxl / nh3 are imported **lazily** inside the methods (house pattern — the app imports
where a wheel is absent) and, crucially, a missing nh3 makes ``_sanitize`` raise, so a render fails
**closed** (returns ``None`` → the edge offers the original), never unsanitised. A broken document
is likewise a ``None`` (offer the original — FR-44), never a raise, never a 500.
"""

from __future__ import annotations

import html

from apx.core.ports.render import RenderedDocument

# The strict allow-list: structural + text-formatting + table tags, and a safe link. Everything else
# — script, every on* handler, javascript:/data: URLs, iframe/object/embed/form/style, and ALL
# images (embedded figures live in the original, offered not inlined this increment) — is stripped.
_ALLOWED_TAGS = frozenset({
    "p", "br", "span", "div", "h1", "h2", "h3", "h4", "h5", "h6",
    "strong", "b", "em", "i", "u", "s", "sub", "sup", "small",
    "blockquote", "pre", "code", "hr",
    "ul", "ol", "li", "dl", "dt", "dd",
    "table", "thead", "tbody", "tfoot", "tr", "td", "th", "caption", "colgroup", "col",
    "a",
})
_ALLOWED_ATTRS = {"a": {"href", "title"}, "th": {"scope"}}
_URL_SCHEMES = frozenset({"http", "https", "mailto"})  # NOT javascript: and NOT data:


def _sanitize(raw_html: str) -> str:
    """Strip ``raw_html`` to the script-free allow-list (nh3/ammonia). Imported lazily so a missing
    wheel makes the caller fail CLOSED (never returns unsanitised). This is the one XSS boundary."""
    import nh3

    return nh3.clean(
        raw_html,
        tags=set(_ALLOWED_TAGS),
        attributes={tag: set(attrs) for tag, attrs in _ALLOWED_ATTRS.items()},
        url_schemes=set(_URL_SCHEMES),
        link_rel="noopener noreferrer",
    )


def _rendered(title: str, raw_html: str, truncated: bool = False) -> RenderedDocument:
    """The ONE construction site for a ``RenderedDocument`` — it sanitises ``raw_html`` before the
    markup can be stored, so no render path can emit unsanitised HTML (a structural property). Every
    renderer routes its output through here. ``title`` is the untrusted pièce filename and is NOT
    sanitised (escaping it here would corrupt a legitimate name like ``Facture & Devis.xlsx``): it
    is text metadata the consumer renders as a text node, never HTML — see ``RenderedDocument``."""
    return RenderedDocument(
        format="html", title=title, html=_sanitize(raw_html), truncated=truncated)


def _suffix(filename: str) -> str:
    return ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""


class HtmlPieceRenderer:
    """Implements the ``PieceRenderer`` port for the office formats rendered server-side (Story
    3.5c-2): ``.docx`` and ``.xlsx``. Any other format → ``None`` (the edge offers the original)."""

    def __init__(self, max_rows: int = 2000, max_cols: int = 64, max_sheets: int = 32) -> None:
        # Grid bounds protect the reader's machine on a dense sheet; a bound hit sets truncated,
        # never a silent drop. Constructor args (injectable at the edge), not hard-coded branches.
        self._max_rows = max_rows
        self._max_cols = max_cols
        self._max_sheets = max_sheets

    def render(self, *, filename: str, data: bytes) -> RenderedDocument | None:
        suffix = _suffix(filename)
        if suffix == ".docx":
            return self._docx(data, filename)
        if suffix == ".xlsx":
            return self._xlsx(data, filename)
        return None  # not a server-rendered office format — the edge offers the original (FR-44)

    def _docx(self, data: bytes, title: str) -> RenderedDocument | None:
        """`.docx` → HTML via mammoth, then sanitised. mammoth escapes text runs; nh3 strips any
        active markup (links, images, styles) mammoth emits. Malformed / empty → None."""
        try:
            import io

            import mammoth

            result = mammoth.convert_to_html(io.BytesIO(data))
            raw = result.value or ""
            if not raw.strip():
                return None  # a .docx with no rendered content — offer the original
            return _rendered(title, raw)
        except Exception:  # noqa: BLE001 — a broken .docx (or missing wheel) is a None, offer original
            return None

    def _xlsx(self, data: bytes, title: str) -> RenderedDocument | None:
        """`.xlsx` → one HTML ``<table>`` per sheet (values, read-only, bounded). Cell values are
        HTML-escaped AND the whole thing is nh3-sanitised (defense in depth). Malformed / empty →
        None. Hitting a row/col/sheet bound sets ``truncated`` (honest, never a silent drop)."""
        try:
            import io

            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception:  # noqa: BLE001 — not a readable .xlsx — offer the original
            return None
        try:
            parts, truncated = self._sheets_html(workbook)
            if not parts:
                return None
            # `_rendered` (nh3) is INSIDE the guard so a sanitiser/nh3 failure fails CLOSED to None
            # (offer the original), symmetric with `_docx` — never an unguarded raise / 500 (AC1).
            return _rendered(title, "".join(parts), truncated=truncated)
        except Exception:  # noqa: BLE001 — a parse/sanitise failure is a None, offer the original
            return None
        finally:
            try:
                workbook.close()  # a read-only workbook holds a file handle until closed
            except Exception:  # noqa: BLE001
                pass

    def _sheets_html(self, workbook: object) -> tuple[list[str], bool]:
        parts: list[str] = []
        truncated = False
        for sheet_index, sheet in enumerate(workbook.worksheets):  # type: ignore[attr-defined]
            if sheet_index >= self._max_sheets:
                truncated = True
                break
            rows_html: list[str] = []
            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_index >= self._max_rows:
                    truncated = True
                    break
                if len(row) > self._max_cols:
                    truncated = True
                cells = "".join(
                    f"<td>{html.escape('' if value is None else str(value))}</td>"
                    for value in row[: self._max_cols])
                rows_html.append(f"<tr>{cells}</tr>")
            caption = html.escape(sheet.title or f"Feuille {sheet_index + 1}")
            parts.append(
                f"<table><caption>{caption}</caption><tbody>{''.join(rows_html)}</tbody></table>")
        return parts, truncated
